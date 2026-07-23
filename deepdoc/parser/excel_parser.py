#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

"""
Excel / CSV 文件解析器

支持 .xlsx、.xls、.csv 等多种电子表格格式的解析。
提供三种输出格式：
1. __call__：以 "字段名：值" 的键值对格式输出，适合后续 LLM 理解
2. html()：输出 HTML <table> 格式，保留表格结构并支持大表分块
3. markdown()：输出 Markdown 表格格式

核心特性：
- 自动识别文件类型（Excel/CSV），通过文件头魔法数字检测
- 多引擎容错：openpyxl → pandas → calamine 链式降级
- 大表优化：二分查找定位实际数据行数，避免遍历空行
- 图片提取：支持提取工作表中的嵌入图片并记录其锚点位置
- 非法字符清理：过滤 Excel 不允许的控制字符
"""

import logging
import re
import sys
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook

from rag.nlp import find_codec
from rag.utils.lazy_image import LazyImage

# 从 openpyxl 复制的非法字符正则：过滤掉 ASCII 控制字符（除了 tab/newline）
ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


class RAGFlowExcelParser:
    """Excel / CSV 电子表格解析器

    支持多种输入格式（.xlsx/.xls/.csv），自动识别文件类型。
    提供键值对、HTML 表格、Markdown 表格三种输出格式。
    """

    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        """加载 Excel/CSV 文件为 openpyxl Workbook 对象

        检测策略：
        1. 读取文件头 4 字节判断文件类型（PK\x03\x04 = xlsx, \xd0\xcf\x11\xe0 = xls）
        2. 如果不匹配，尝试作为 CSV 解析
        3. 如果是 Excel 格式，依次降级尝试：openpyxl → pandas → calamine

        Args:
            file_like_object: 文件路径、bytes 或 BytesIO 对象

        Returns:
            openpyxl Workbook 对象

        Raises:
            Exception: 所有解析方式均失败时抛出
        """
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)

        # 读取文件头 4 字节判断类型
        file_like_object.seek(0)
        file_head = file_like_object.read(4)
        file_like_object.seek(0)

        # 非 Excel 文件头 → 尝试 CSV
        if not (file_head.startswith(b"PK\x03\x04") or file_head.startswith(b"\xd0\xcf\x11\xe0")):
            logging.info("Not an Excel file, converting CSV to Excel Workbook")
            try:
                file_like_object.seek(0)
                df = pd.read_csv(file_like_object, on_bad_lines='skip')
                return RAGFlowExcelParser._dataframe_to_workbook(df)
            except Exception as e_csv:
                raise Exception(f"Failed to parse CSV and convert to Excel Workbook: {e_csv}")

        # Excel 格式：openpyxl → pandas → calamine 链式降级
        try:
            return load_workbook(file_like_object, data_only=True)
        except Exception as e:
            logging.info(f"openpyxl load error: {e}, try pandas instead")
            try:
                file_like_object.seek(0)
                try:
                    dfs = pd.read_excel(file_like_object, sheet_name=None)
                    return RAGFlowExcelParser._dataframe_to_workbook(dfs)
                except Exception as ex:
                    logging.info(f"pandas with default engine load error: {ex}, try calamine instead")
                    file_like_object.seek(0)
                    df = pd.read_excel(file_like_object, engine="calamine")
                    return RAGFlowExcelParser._dataframe_to_workbook(df)
            except Exception as e_pandas:
                raise Exception(f"pandas.read_excel error: {e_pandas}, original openpyxl error: {e}")

    @staticmethod
    def _clean_dataframe(df: pd.DataFrame):
        """清理 DataFrame 中的非法字符

        Excel 不允许部分 ASCII 控制字符，需要替换为空格。
        """
        def clean_string(s):
            if isinstance(s, str):
                return ILLEGAL_CHARACTERS_RE.sub(" ", s)
            return s

        return df.apply(lambda col: col.map(clean_string))

    @staticmethod
    def _fill_worksheet_from_dataframe(ws, df: pd.DataFrame):
        """将 DataFrame 数据填充到 openpyxl Worksheet 中"""
        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)
        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

    @staticmethod
    def _dataframe_to_workbook(df):
        """将 DataFrame 或 {sheet_name: DataFrame} 字典转为 Workbook"""
        if isinstance(df, dict) and len(df) > 1:
            return RAGFlowExcelParser._dataframes_to_workbook(df)

        df = RAGFlowExcelParser._clean_dataframe(df)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        RAGFlowExcelParser._fill_worksheet_from_dataframe(ws, df)
        return wb

    @staticmethod
    def _dataframes_to_workbook(dfs: dict):
        """将多工作表字典转为 Workbook（每个 key 为一个 sheet）"""
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        for sheet_name, df in dfs.items():
            df = RAGFlowExcelParser._clean_dataframe(df)
            ws = wb.create_sheet(title=sheet_name)
            RAGFlowExcelParser._fill_worksheet_from_dataframe(ws, df)
        return wb

    @staticmethod
    def _extract_images_from_worksheet(ws, sheetname=None):
        """从工作表中提取嵌入的图片及其锚点位置

        通过 openpyxl 的 _images 属性获取图片列表，
        解析每个图片的锚点（anchor）确定其所在的行列位置。

        Returns:
            dict 列表，每项包含：sheet, image(LazyImage), row_from, col_from,
            row_to, col_to, span_type(single_cell/multi_cell)
        """
        images = getattr(ws, "_images", [])
        if not images:
            return []

        raw_items = []

        for img in images:
            try:
                img_bytes = img._data()
                lazy_img = LazyImage([img_bytes])

                anchor = img.anchor
                if hasattr(anchor, "_from") and hasattr(anchor, "_to"):
                    r1, c1 = anchor._from.row + 1, anchor._from.col + 1
                    r2, c2 = anchor._to.row + 1, anchor._to.col + 1
                    if r1 == r2 and c1 == c2:
                        span = "single_cell"
                    else:
                        span = "multi_cell"
                else:
                    r1, c1 = anchor._from.row + 1, anchor._from.col + 1
                    r2, c2 = r1, c1
                    span = "single_cell"

                item = {
                    "sheet": sheetname or ws.title,
                    "image": lazy_img,
                    "image_description": "",
                    "row_from": r1,
                    "col_from": c1,
                    "row_to": r2,
                    "col_to": c2,
                    "span_type": span,
                }
                raw_items.append(item)
            except Exception:
                continue
        return raw_items

    @staticmethod
    def _get_actual_row_count(ws):
        """通过二分查找高效定位工作表中的实际数据行数

        openpyxl 的 ws.max_row 可能远大于实际有数据的行数
        （特别是被编辑过的文件保留了已删除行的格式信息）。
        此方法使用二分查找 + 前后各 10 行的采样来快速估计真实行数。

        Args:
            ws: openpyxl Worksheet 对象

        Returns:
            估计的实际数据行数（0 表示没有数据）
        """
        max_row = ws.max_row
        if not max_row:
            return 0
        if max_row <= 10000:
            return max_row

        max_col = min(ws.max_column or 1, 50)

        def row_has_data(row_idx):
            """检查指定行是否有任何非空单元格"""
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    return True
            return False

        # 如果前 100 行都没有数据，认为是空表
        if not any(row_has_data(i) for i in range(1, min(101, max_row + 1))):
            return 0

        # 二分查找有数据的最后一行
        left, right = 1, max_row
        last_data_row = 1

        while left <= right:
            mid = (left + right) // 2
            found = False
            for r in range(mid, min(mid + 10, max_row + 1)):
                if row_has_data(r):
                    found = True
                    last_data_row = max(last_data_row, r)
                    break
            if found:
                left = mid + 1
            else:
                right = mid - 1

        # 在找到的最后位置前后 500 行做最终确认
        for r in range(last_data_row, min(last_data_row + 500, max_row + 1)):
            if row_has_data(r):
                last_data_row = r

        return last_data_row

    @staticmethod
    def _get_rows_limited(ws):
        """获取工作表的行数据（仅包含实际有数据的行）"""
        actual_rows = RAGFlowExcelParser._get_actual_row_count(ws)
        if actual_rows == 0:
            return []
        return list(ws.iter_rows(min_row=1, max_row=actual_rows))

    def html(self, fnm, chunk_rows=256):
        """将电子表格转换为 HTML <table> 格式

        大表自动按 chunk_rows 行数分块，每个块独立为一段。

        Args:
            fnm: 文件路径或二进制内容
            chunk_rows: 每个 HTML 表格块的最大数据行数

        Returns:
            HTML 表格字符串列表
        """
        from html import escape

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)
        tb_chunks = []

        def _fmt(v):
            if v is None:
                return ""
            return str(v).strip()

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            try:
                rows = RAGFlowExcelParser._get_rows_limited(ws)
            except Exception as e:
                logging.warning(f"Skip sheet '{sheetname}' due to rows access error: {e}")
                continue
            if not rows:
                continue

            # 构建表头行
            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{escape(_fmt(t.value))}</th>"
            tb_rows_0 += "</tr>"

            # 按 chunk_rows 分块输出数据行
            n_data_rows = len(rows) - 1
            for chunk_i in range((n_data_rows + chunk_rows - 1) // chunk_rows):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(rows[1 + chunk_i * chunk_rows : min(1 + (chunk_i + 1) * chunk_rows, len(rows))]):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{escape(_fmt(c.value))}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    def markdown(self, fnm):
        """将电子表格转换为 Markdown 表格格式

        对 CSV 文件有额外的兼容处理：如果 Excel 解析失败，自动降级为 CSV 解析。
        """
        import pandas as pd

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        try:
            file_like_object.seek(0)
            df = pd.read_excel(file_like_object)
        except Exception as e:
            logging.warning(f"Parse spreadsheet error: {e}, trying to interpret as CSV file")
            file_like_object.seek(0)
            df = pd.read_csv(file_like_object, on_bad_lines='skip')
        df = df.replace(r"^\s*$", "", regex=True)
        return df.to_markdown(index=False)

    def __call__(self, fnm):
        """将电子表格解析为键值对格式的文本行列表

        每行的输出格式：'列1名：值1; 列2名：值2 ——工作表名'
        只输出有值的列，空值跳过。如果工作表名不包含 "sheet"，
        则在行末追加工作表名作为来源标记。

        Args:
            fnm: 文件路径或二进制内容

        Returns:
            键值对格式的文本行列表
        """
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)

        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            try:
                rows = RAGFlowExcelParser._get_rows_limited(ws)
            except Exception as e:
                logging.warning(f"Skip sheet '{sheetname}' due to rows access error: {e}")
                continue
            if not rows:
                continue
            ti = list(rows[0])  # 表头行
            for r in list(rows[1:]):
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    fields.append(t)
                if not fields:
                    continue
                line = "; ".join(fields)
                # 如果工作表名不包含 "sheet"，则追加工作表名作为上下文
                if sheetname.lower().find("sheet") < 0:
                    line += " ——" + sheetname
                res.append(line)
        return res

    @staticmethod
    def row_number(fnm, binary):
        """获取电子表格的总行数（用于进度估计）

        Args:
            fnm: 文件名
            binary: 文件二进制内容

        Returns:
            总行数（所有工作表的数据行之和）
        """
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = RAGFlowExcelParser._load_excel_to_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                try:
                    ws = wb[sheetname]
                    total += RAGFlowExcelParser._get_actual_row_count(ws)
                except Exception as e:
                    logging.warning(f"Skip sheet '{sheetname}' due to rows access error: {e}")
                    continue
            return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    psr(sys.argv[1])
